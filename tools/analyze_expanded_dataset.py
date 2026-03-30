#!/usr/bin/env python3
"""Generate reproducibility and robustness outputs for the footwear manuscript datasets."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import stats


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DATASET = ROOT / "expanded_dataset_40_models.csv"
DEFAULT_PRODUCT_SOURCES = ROOT / "product_source_appendix_40_models.csv"
DEFAULT_OUTDIR = ROOT / "supplementary_outputs"
DEFAULT_WORKBOOK = ROOT / "Supplementary_Materials_Statistical_Strengthening.xlsx"

LEATHER_TYPES = {"Full Leather", "Partial Leather", "Synthetic Leather"}
BOOTSTRAP_SEED = 42
BOOTSTRAP_N = 10_000


@dataclass
class OLSResult:
    coefficients: pd.DataFrame
    r2: float
    adjusted_r2: float
    f_stat: float
    f_pvalue: float
    residual_df: int


def ols(y: pd.Series, X: pd.DataFrame) -> OLSResult:
    yv = y.astype(float).to_numpy()
    Xv = X.astype(float).to_numpy()
    n, k = Xv.shape
    xtx_inv = np.linalg.pinv(Xv.T @ Xv)
    beta = xtx_inv @ (Xv.T @ yv)
    fitted = Xv @ beta
    resid = yv - fitted
    rank = int(np.linalg.matrix_rank(Xv))
    residual_df = max(n - rank, 1)
    sse = float(resid.T @ resid)
    sigma2 = sse / residual_df
    se = np.sqrt(np.diag(xtx_inv) * sigma2)
    tvals = np.divide(beta, se, out=np.full_like(beta, np.nan), where=se > 0)
    pvals = 2 * (1 - stats.t.cdf(np.abs(tvals), residual_df))
    sst = float(((yv - yv.mean()) ** 2).sum())
    r2 = 1 - sse / sst if sst else 0.0
    adjusted_r2 = 1 - (1 - r2) * (n - 1) / residual_df if residual_df else r2
    explained_df = max(rank - 1, 1)
    ssr = float(((fitted - yv.mean()) ** 2).sum())
    f_stat = (ssr / explained_df) / (sse / residual_df) if sse > 0 else np.inf
    f_pvalue = 1 - stats.f.cdf(f_stat, explained_df, residual_df) if np.isfinite(f_stat) else 0.0
    return OLSResult(
        coefficients=pd.DataFrame(
            {
                "coef": beta,
                "se": se,
                "t": tvals,
                "p": pvals,
            },
            index=X.columns,
        ),
        r2=r2,
        adjusted_r2=adjusted_r2,
        f_stat=f_stat,
        f_pvalue=f_pvalue,
        residual_df=residual_df,
    )


def hc3_statistics(y: pd.Series, X: pd.DataFrame) -> dict[str, np.ndarray | float]:
    yv = y.astype(float).to_numpy()
    Xv = X.astype(float).to_numpy()
    xtx_inv = np.linalg.pinv(Xv.T @ Xv)
    beta = xtx_inv @ (Xv.T @ yv)
    fitted = Xv @ beta
    resid = yv - fitted
    rank = int(np.linalg.matrix_rank(Xv))
    residual_df = max(len(yv) - rank, 1)
    leverage = np.sum(Xv * (Xv @ xtx_inv), axis=1)
    scale = np.divide(resid, 1 - leverage, out=np.zeros_like(resid), where=(1 - leverage) != 0) ** 2
    meat = np.zeros((Xv.shape[1], Xv.shape[1]))
    for idx in range(len(yv)):
        xi = Xv[idx : idx + 1].T
        meat += scale[idx] * (xi @ xi.T)
    vcov = xtx_inv @ meat @ xtx_inv
    se = np.sqrt(np.diag(vcov))
    tvals = np.divide(beta, se, out=np.full_like(beta, np.nan), where=se > 0)
    pvals = 2 * (1 - stats.t.cdf(np.abs(tvals), residual_df))
    return {
        "beta": beta,
        "se": se,
        "t": tvals,
        "p": pvals,
        "resid": resid,
        "fitted": fitted,
        "leverage": leverage,
        "residual_df": residual_df,
    }


def breusch_pagan(resid: np.ndarray, X: pd.DataFrame) -> tuple[float, float]:
    auxiliary = ols(pd.Series(np.asarray(resid) ** 2), X)
    yv = np.asarray(resid) ** 2
    fitted = X.astype(float).to_numpy() @ auxiliary.coefficients["coef"].to_numpy()
    sst = float(((yv - yv.mean()) ** 2).sum())
    sse = float(((yv - fitted) ** 2).sum())
    r2 = 1 - sse / sst if sst > 0 else 0.0
    lm = len(yv) * r2
    pvalue = 1 - stats.chi2.cdf(lm, X.shape[1] - 1)
    return float(lm), float(pvalue)


def autosize_worksheet(ws) -> None:
    wrap = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = wrap
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in ws[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)


def dataframe_to_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    ws = workbook.create_sheet(title[:31])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    ws.append(list(frame.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in frame.itertuples(index=False, name=None):
        ws.append(list(row))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_worksheet(ws)


def write_csv(path: Path, rows: Iterable[dict]) -> pd.DataFrame:
    frame = pd.DataFrame(list(rows))
    frame.to_csv(path, index=False)
    return frame


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dataset", type=Path, default=DEFAULT_DATASET, help="CSV dataset to analyze")
    parser.add_argument(
        "--product-sources",
        type=Path,
        default=DEFAULT_PRODUCT_SOURCES,
        help="Product-source appendix CSV aligned to the dataset",
    )
    parser.add_argument(
        "--outdir",
        type=Path,
        default=DEFAULT_OUTDIR,
        help="Directory for generated supplementary CSV outputs",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="Workbook path for generated Excel outputs",
    )
    return parser.parse_args()


def load_dataset(dataset_path: Path) -> pd.DataFrame:
    df = pd.read_csv(dataset_path)
    df["Is_Leather"] = df["Upper_Material_Type"].isin(LEATHER_TYPES).astype(int)
    df["Leather_Group"] = np.where(df["Is_Leather"] == 1, "Leather-associated", "Non-leather")
    df["Annualized_CO2e"] = df["CO2e_Total"] / df["Lifespan_Years"]
    return df


def summary_statistics(df: pd.DataFrame) -> pd.DataFrame:
    series = df["CO2e_Total"]
    rows = [
        {"Statistic": "n", "Value": len(series)},
        {"Statistic": "Mean", "Value": round(series.mean(), 3)},
        {"Statistic": "Median", "Value": round(series.median(), 3)},
        {"Statistic": "Std_Dev", "Value": round(series.std(ddof=1), 3)},
        {"Statistic": "Min", "Value": round(series.min(), 3)},
        {"Statistic": "Max", "Value": round(series.max(), 3)},
        {"Statistic": "Skewness", "Value": round(series.skew(), 3)},
        {"Statistic": "Excess_Kurtosis", "Value": round(series.kurt(), 3)},
        {"Statistic": "CV", "Value": round(series.std(ddof=1) / series.mean(), 3)},
        {
            "Statistic": "Material_Share",
            "Value": round(df["CO2e_Materials"].sum() / df["CO2e_Total"].sum(), 3),
        },
    ]
    return pd.DataFrame(rows)


def category_statistics(df: pd.DataFrame) -> pd.DataFrame:
    frame = (
        df.groupby("Category")["CO2e_Total"]
        .agg(n="count", Mean="mean", SD="std", Min="min", Max="max")
        .reset_index()
    )
    for col in ["Mean", "SD", "Min", "Max"]:
        frame[col] = frame[col].round(3)
    return frame


def material_statistics(df: pd.DataFrame) -> pd.DataFrame:
    frame = (
        df.groupby("Upper_Material_Type")
        .agg(
            n=("CO2e_Total", "count"),
            Mean_CO2e=("CO2e_Total", "mean"),
            SD_CO2e=("CO2e_Total", "std"),
            Mean_Water_L=("Water_L", "mean"),
            SD_Water_L=("Water_L", "std"),
        )
        .reset_index()
        .sort_values(["Mean_CO2e", "Upper_Material_Type"], ascending=[False, True])
    )
    for col in ["Mean_CO2e", "SD_CO2e", "Mean_Water_L", "SD_Water_L"]:
        frame[col] = frame[col].round(3)
    return frame


def reproducibility_rules(dataset_name: str, outdir_name: str) -> pd.DataFrame:
    rows = [
        {
            "Item": "Functional unit",
            "Rule": "Primary comparisons use one pair of footwear; service-life-normalized sensitivity is reported separately as Annualized_CO2e",
            "Source_or_File": f"Research_Paper_MDPI_Strengthened.md; {dataset_name}",
        },
        {
            "Item": "System boundary",
            "Rule": "Cradle-to-grave screening model informed by ISO 14040/14044 principles, with materials, manufacturing, transport, packaging, and end-of-life stages",
            "Source_or_File": "Research_Paper_MDPI_Strengthened.md",
        },
        {
            "Item": "Mass basis",
            "Rule": "Mass_kg is a stored reference mass proxy in kilograms; public-disclosure reconciliation suggests it behaves closer to single-shoe scale than doubled pair scale, exact size-specific variation is not modeled separately, and the released 40-row carbon totals are retained unchanged unless a future mass-calibrated model revision is built explicitly",
            "Source_or_File": f"{dataset_name}; {outdir_name}/S16_Assumption_Rules.csv",
        },
        {
            "Item": "Leather-associated coding",
            "Rule": "Is_Leather = 1 when Upper_Material_Type is Full Leather, Partial Leather, or Synthetic Leather",
            "Source_or_File": dataset_name,
        },
        {
            "Item": "Category coding",
            "Rule": "Running, Basketball, Casual, Soccer categories carried directly from the released dataset",
            "Source_or_File": dataset_name,
        },
        {
            "Item": "Price variable",
            "Rule": "Price_USD is the list price used in all reported regressions and correlations",
            "Source_or_File": dataset_name,
        },
        {
            "Item": "Phase allocation formula",
            "Rule": "CO2e_Total = CO2e_Materials + CO2e_Manufacturing + CO2e_Transport + CO2e_Packaging + CO2e_EOL",
            "Source_or_File": dataset_name,
        },
        {
            "Item": "Phase multipliers",
            "Rule": "Manufacturing ~= 25% of materials; transport ~= 15%; packaging ~= 5%; end-of-life ~= 2%; total ~= 1.47 x materials-stage estimate before rounding",
            "Source_or_File": f"{dataset_name}; {outdir_name}/S14_Inventory_Calculation.csv",
        },
        {
            "Item": "Durability scenario",
            "Rule": "Annualized_CO2e = CO2e_Total / Lifespan_Years; product-specific lifespan buckets and rationales are released in the supplementary rules",
            "Source_or_File": f"{dataset_name}; {outdir_name}/S16_Assumption_Rules.csv",
        },
        {
            "Item": "Sampling sensitivity",
            "Rule": "Robustness checks include leave-one-category-out regressions and balanced-material bootstrap resampling",
            "Source_or_File": f"{outdir_name}/S17_Sampling_Sensitivity.csv",
        },
        {
            "Item": "Regression diagnostics",
            "Rule": "Full coefficient tables include HC3 robust standard errors; residual normality, heteroskedasticity, leverage, and Cook's distance are summarized separately",
            "Source_or_File": f"{outdir_name}/S8_Regressions.csv; {outdir_name}/S18_Regression_Diagnostics.csv",
        },
        {
            "Item": "Water and ecotoxicity",
            "Rule": "Comparative screening indicators are used exactly as released in the dataset and analyzed as product-level outcomes",
            "Source_or_File": dataset_name,
        },
        {
            "Item": "Reproducibility code",
            "Rule": "Run python3 tools/analyze_expanded_dataset.py then python3 tools/regenerate_artifacts.py",
            "Source_or_File": "tools/analyze_expanded_dataset.py; tools/regenerate_artifacts.py",
        },
    ]
    return pd.DataFrame(rows)


def correlations_and_tests(df: pd.DataFrame) -> pd.DataFrame:
    pearson = stats.pearsonr(df["Price_USD"], df["CO2e_Total"])
    spearman = stats.spearmanr(df["Price_USD"], df["CO2e_Total"])
    leather = df.loc[df["Is_Leather"] == 1, "Water_L"]
    nonleather = df.loc[df["Is_Leather"] == 0, "Water_L"]
    welch = stats.ttest_ind(leather, nonleather, equal_var=False)
    pooled = np.sqrt(
        (
            (len(leather) - 1) * leather.var(ddof=1)
            + (len(nonleather) - 1) * nonleather.var(ddof=1)
        )
        / (len(leather) + len(nonleather) - 2)
    )
    d = (leather.mean() - nonleather.mean()) / pooled
    return pd.DataFrame(
        [
            {"Test": "Pearson Price vs CO2e", "Statistic": round(pearson.statistic, 4), "P_Value": pearson.pvalue},
            {"Test": "Spearman Price vs CO2e", "Statistic": round(spearman.statistic, 4), "P_Value": spearman.pvalue},
            {"Test": "Welch t Water Leather vs Non-leather", "Statistic": round(welch.statistic, 4), "P_Value": welch.pvalue},
            {"Test": "Cohen d Water Leather vs Non-leather", "Statistic": round(d, 4), "P_Value": np.nan},
        ]
    )


def anova_and_tukey(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    categories = sorted(df["Category"].unique())
    groups = [df.loc[df["Category"] == category, "CO2e_Total"].to_numpy() for category in categories]
    anova = stats.f_oneway(*groups)
    grand_mean = df["CO2e_Total"].mean()
    ss_between = sum(len(group) * (group.mean() - grand_mean) ** 2 for group in groups)
    ss_total = ((df["CO2e_Total"] - grand_mean) ** 2).sum()
    eta2 = ss_between / ss_total
    anova_frame = pd.DataFrame(
        [
            {
                "Statistic": "F",
                "Value": round(anova.statistic, 4),
                "Df1": len(categories) - 1,
                "Df2": len(df) - len(categories),
                "P_Value": anova.pvalue,
                "Eta_Squared": round(eta2, 4),
            }
        ]
    )

    tukey = stats.tukey_hsd(*groups)
    rows = []
    for i, left in enumerate(categories):
        for j, right in enumerate(categories):
            if i >= j:
                continue
            rows.append(
                {
                    "Comparison": f"{left} vs {right}",
                    "Mean_Difference": round(float(tukey.statistic[i, j]), 4),
                    "P_Value": float(tukey.pvalue[i, j]),
                    "CI_Low": round(float(tukey.confidence_interval().low[i, j]), 4),
                    "CI_High": round(float(tukey.confidence_interval().high[i, j]), 4),
                }
            )
    return anova_frame, pd.DataFrame(rows)


def regression_models(df: pd.DataFrame) -> pd.DataFrame:
    models = []

    def add_model(name: str, y: pd.Series, X: pd.DataFrame) -> None:
        result = ols(y, X)
        hc3 = hc3_statistics(y, X)
        crit = stats.t.ppf(0.975, result.residual_df)
        for term, row in result.coefficients.iterrows():
            idx = list(X.columns).index(term)
            coef = float(row["coef"])
            se = float(row["se"])
            hc3_se = float(hc3["se"][idx])
            models.append(
                {
                    "Model": name,
                    "Term": term,
                    "Estimate": round(coef, 6),
                    "SE": round(se, 6),
                    "t": round(float(row["t"]), 6),
                    "P_Value": float(row["p"]),
                    "CI95_Low": round(coef - crit * se, 6),
                    "CI95_High": round(coef + crit * se, 6),
                    "HC3_SE": round(hc3_se, 6),
                    "HC3_t": round(float(hc3["t"][idx]), 6),
                    "HC3_P_Value": float(hc3["p"][idx]),
                    "HC3_CI95_Low": round(coef - crit * hc3_se, 6),
                    "HC3_CI95_High": round(coef + crit * hc3_se, 6),
                    "R2": round(result.r2, 6),
                    "Adj_R2": round(result.adjusted_r2, 6),
                    "F_Statistic": round(result.f_stat, 6),
                    "F_PValue": float(result.f_pvalue),
                }
            )

    baseline = pd.DataFrame(
        {
            "Intercept": 1.0,
            "Price_USD": df["Price_USD"],
            "Is_Leather": df["Is_Leather"],
        }
    )
    add_model("Baseline", df["CO2e_Total"], baseline)

    category_dummies = pd.get_dummies(df["Category"], drop_first=True)
    controlled = pd.concat([baseline, category_dummies], axis=1)
    add_model("Category_Adjusted", df["CO2e_Total"], controlled)

    log_price = pd.DataFrame(
        {
            "Intercept": 1.0,
            "log_Price_USD": np.log(df["Price_USD"]),
            "Is_Leather": df["Is_Leather"],
        }
    )
    add_model("Log_Price", df["CO2e_Total"], log_price)

    trimmed = df.drop(df["CO2e_Total"].idxmax()).reset_index(drop=True)
    trimmed_X = pd.DataFrame(
        {
            "Intercept": 1.0,
            "Price_USD": trimmed["Price_USD"],
            "Is_Leather": trimmed["Is_Leather"],
        }
    )
    add_model("Max_Outlier_Removed", trimmed["CO2e_Total"], trimmed_X)

    return pd.DataFrame(models)


def regression_diagnostics(df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    specs = {
        "Baseline": (
            df["CO2e_Total"],
            pd.DataFrame({"Intercept": 1.0, "Price_USD": df["Price_USD"], "Is_Leather": df["Is_Leather"]}),
        ),
        "Category_Adjusted": (
            df["CO2e_Total"],
            pd.concat(
                [
                    pd.DataFrame({"Intercept": 1.0, "Price_USD": df["Price_USD"], "Is_Leather": df["Is_Leather"]}),
                    pd.get_dummies(df["Category"], drop_first=True).astype(int),
                ],
                axis=1,
            ),
        ),
        "Annualized_Baseline": (
            df["Annualized_CO2e"],
            pd.DataFrame({"Intercept": 1.0, "Price_USD": df["Price_USD"], "Is_Leather": df["Is_Leather"]}),
        ),
    }

    for name, (y, X) in specs.items():
        result = ols(y, X)
        hc3 = hc3_statistics(y, X)
        lm, bp_p = breusch_pagan(np.asarray(hc3["resid"]), X)
        sigma = np.sqrt(float((np.asarray(hc3["resid"]) ** 2).sum()) / result.residual_df)
        leverage = np.asarray(hc3["leverage"])
        studentized = np.divide(
            np.asarray(hc3["resid"]),
            sigma * np.sqrt(np.maximum(1 - leverage, 1e-9)),
            out=np.zeros_like(np.asarray(hc3["resid"])),
            where=sigma > 0,
        )
        cooks = (studentized**2 * leverage) / (X.shape[1] * np.maximum(1 - leverage, 1e-9))
        max_idx = int(np.argmax(cooks))
        leather_idx = list(X.columns).index("Is_Leather") if "Is_Leather" in X.columns else None
        price_idx = list(X.columns).index("Price_USD") if "Price_USD" in X.columns else None
        rows.append(
            {
                "Model": name,
                "Outcome": y.name if y.name else "Outcome",
                "N": len(y),
                "Residual_Shapiro_P": float(stats.shapiro(np.asarray(hc3["resid"])).pvalue),
                "Breusch_Pagan_LM": round(lm, 6),
                "Breusch_Pagan_P": bp_p,
                "Max_Leverage": round(float(leverage.max()), 6),
                "Max_Abs_Studentized_Residual": round(float(np.abs(studentized).max()), 6),
                "Max_Cooks_D": round(float(cooks.max()), 6),
                "Most_Influential_Model": df.iloc[max_idx]["Model"],
                "Leather_Estimate": round(float(result.coefficients.loc["Is_Leather", "coef"]), 6),
                "Leather_HC3_SE": round(float(hc3["se"][leather_idx]), 6) if leather_idx is not None else np.nan,
                "Leather_HC3_P_Value": float(hc3["p"][leather_idx]) if leather_idx is not None else np.nan,
                "Price_HC3_SE": round(float(hc3["se"][price_idx]), 6) if price_idx is not None else np.nan,
                "Price_HC3_P_Value": float(hc3["p"][price_idx]) if price_idx is not None else np.nan,
                "R2": round(result.r2, 6),
            }
        )

    return pd.DataFrame(rows)


def leave_one_out(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    for idx, model_name in enumerate(df["Model"]):
        subset = df.drop(df.index[idx]).reset_index(drop=True)
        X = pd.DataFrame(
            {
                "Intercept": 1.0,
                "Price_USD": subset["Price_USD"],
                "Is_Leather": subset["Is_Leather"],
            }
        )
        result = ols(subset["CO2e_Total"], X)
        leather_row = result.coefficients.loc["Is_Leather"]
        rows.append(
            {
                "Left_Out": model_name,
                "Leather_Estimate": round(float(leather_row["coef"]), 6),
                "Leather_SE": round(float(leather_row["se"]), 6),
                "Leather_t": round(float(leather_row["t"]), 6),
                "Leather_P_Value": float(leather_row["p"]),
                "R2": round(result.r2, 6),
            }
        )
    full = pd.DataFrame(rows).sort_values("Leather_Estimate").reset_index(drop=True)
    summary = pd.DataFrame(
        [
            {
                "Statistic": "Min Leather Estimate",
                "Value": round(full["Leather_Estimate"].min(), 6),
            },
            {
                "Statistic": "Median Leather Estimate",
                "Value": round(full["Leather_Estimate"].median(), 6),
            },
            {
                "Statistic": "Max Leather Estimate",
                "Value": round(full["Leather_Estimate"].max(), 6),
            },
            {
                "Statistic": "Max Leather P_Value",
                "Value": float(full["Leather_P_Value"].max()),
            },
            {
                "Statistic": "Min R2",
                "Value": round(full["R2"].min(), 6),
            },
            {
                "Statistic": "Max R2",
                "Value": round(full["R2"].max(), 6),
            },
        ]
    )
    return full, summary


def bootstrap_intervals(df: pd.DataFrame) -> pd.DataFrame:
    rng = np.random.default_rng(BOOTSTRAP_SEED)
    categories = sorted(df["Category"].unique())
    overall_mean = []
    overall_median = []
    overall_sd = []
    overall_skew = []
    corr = []
    leather_coef = []
    category_means = {category: [] for category in categories}

    for _ in range(BOOTSTRAP_N):
        sample = df.iloc[rng.integers(0, len(df), len(df))].reset_index(drop=True)
        values = sample["CO2e_Total"]
        overall_mean.append(values.mean())
        overall_median.append(values.median())
        overall_sd.append(values.std(ddof=1))
        overall_skew.append(values.skew())
        corr.append(stats.pearsonr(sample["Price_USD"], sample["CO2e_Total"]).statistic)
        X = pd.DataFrame(
            {
                "Intercept": 1.0,
                "Price_USD": sample["Price_USD"],
                "Is_Leather": sample["Is_Leather"],
            }
        )
        leather_coef.append(float(ols(sample["CO2e_Total"], X).coefficients.loc["Is_Leather", "coef"]))
        grouped = sample.groupby("Category")["CO2e_Total"].mean()
        for category in categories:
            category_means[category].append(float(grouped.get(category, np.nan)))

    rows = []
    for label, values in [
        ("Mean CO2e", overall_mean),
        ("Median CO2e", overall_median),
        ("SD CO2e", overall_sd),
        ("Skewness", overall_skew),
        ("Pearson r Price vs CO2e", corr),
        ("Leather coefficient", leather_coef),
    ]:
        series = pd.Series(values).dropna()
        rows.append(
            {
                "Statistic": label,
                "Bootstrap_Mean": round(series.mean(), 6),
                "Bootstrap_SE": round(series.std(ddof=1), 6),
                "CI_2.5": round(series.quantile(0.025), 6),
                "CI_50": round(series.quantile(0.5), 6),
                "CI_97.5": round(series.quantile(0.975), 6),
            }
        )

    for category in categories:
        series = pd.Series(category_means[category]).dropna()
        rows.append(
            {
                "Statistic": f"Category mean {category}",
                "Bootstrap_Mean": round(series.mean(), 6),
                "Bootstrap_SE": round(series.std(ddof=1), 6),
                "CI_2.5": round(series.quantile(0.025), 6),
                "CI_50": round(series.quantile(0.5), 6),
                "CI_97.5": round(series.quantile(0.975), 6),
            }
        )

    return pd.DataFrame(rows)


def classify_mass_source(mass_basis: str) -> str:
    text = str(mass_basis).lower()
    if "single-shoe weight" in text or "single shoe weight" in text:
        return "single_shoe_disclosure_converted_to_pair"
    if "public weight disclosure" in text or "exact weight disclosure" in text:
        return "direct_public_pair_weight"
    if "anchored to comparable" in text or "inferred from comparable" in text:
        return "architecture_matched_inference"
    return "screening_assumption"


def inventory_calculation_sheet(df: pd.DataFrame, product_sources: pd.DataFrame) -> pd.DataFrame:
    frame = df[
        [
            "Model",
            "Brand",
            "Category",
            "Upper_Material",
            "Upper_Material_Type",
            "Midsole",
            "Outsole",
            "Mass_kg",
            "Price_USD",
            "Lifespan_Years",
            "CO2e_Materials",
            "CO2e_Manufacturing",
            "CO2e_Transport",
            "CO2e_Packaging",
            "CO2e_EOL",
            "CO2e_Total",
            "Annualized_CO2e",
        ]
    ].copy()
    source_cols = product_sources[
        [
            "Model",
            "Source_Type",
            "Product_Source_URL",
            "Mass_Basis",
            "Lifespan_Basis",
            "Traceability_Note",
        ]
    ].copy()
    source_cols["Mass_Source_Type"] = source_cols["Mass_Basis"].map(classify_mass_source)
    frame = frame.merge(source_cols, on="Model", how="left")
    frame["Material_Architecture"] = (
        frame["Upper_Material"] + " / " + frame["Midsole"] + " / " + frame["Outsole"]
    )
    frame["Architecture_Code"] = (
        frame["Category"] + " | " + frame["Upper_Material_Type"] + " | " + frame["Midsole"] + " | " + frame["Outsole"]
    )
    frame["Manufacturing_to_Materials"] = (frame["CO2e_Manufacturing"] / frame["CO2e_Materials"]).round(3)
    frame["Transport_to_Materials"] = (frame["CO2e_Transport"] / frame["CO2e_Materials"]).round(3)
    frame["Packaging_to_Materials"] = (frame["CO2e_Packaging"] / frame["CO2e_Materials"]).round(3)
    frame["EOL_to_Materials"] = (frame["CO2e_EOL"] / frame["CO2e_Materials"]).round(3)
    frame["Implied_Total_to_Materials"] = (frame["CO2e_Total"] / frame["CO2e_Materials"]).round(3)
    frame["Reconstructed_Total"] = (
        frame["CO2e_Materials"]
        + frame["CO2e_Manufacturing"]
        + frame["CO2e_Transport"]
        + frame["CO2e_Packaging"]
        + frame["CO2e_EOL"]
    ).round(6)
    frame["Reconstruction_Error"] = (frame["Reconstructed_Total"] - frame["CO2e_Total"]).round(6)
    frame["Materials_CO2e_per_kg_Stored_Mass"] = (frame["CO2e_Materials"] / frame["Mass_kg"]).round(3)
    frame["Total_CO2e_per_kg_Stored_Mass"] = (frame["CO2e_Total"] / frame["Mass_kg"]).round(3)
    frame["Mass_Field_Interpretation"] = (
        "Mass_kg is treated as a stored reference mass proxy; the public-disclosure validation subset suggests the released values behave closer to single-shoe-scale official weights than to doubled pair-converted masses."
    )
    frame["Downstream_Adder_Logic"] = "Manufacturing ~= 0.25M; Transport ~= 0.15M; Packaging ~= 0.05M; End-of-life ~= 0.02M"
    frame["Calculation_Summary"] = (
        "Total ~= Materials + 0.25M + 0.15M + 0.05M + 0.02M; released phase values are rounded product-level screening estimates."
    )
    return frame


def worked_examples(df: pd.DataFrame, product_sources: pd.DataFrame) -> pd.DataFrame:
    rows = []
    ordered_categories = ["Running", "Basketball", "Casual", "Soccer"]
    sources = product_sources.set_index("Model")
    for category in ordered_categories:
        subset = df.loc[df["Category"] == category].copy()
        median = subset["CO2e_Total"].median()
        selected = subset.iloc[(subset["CO2e_Total"] - median).abs().argsort().iloc[0]]
        material = float(selected["CO2e_Materials"])
        source_row = sources.loc[selected["Model"]]
        rows.append(
            {
                "Category": category,
                "Representative_Model": selected["Model"],
                "Selection_Basis": "Nearest category median CO2e model",
                "Upper_Material": selected["Upper_Material"],
                "Upper_Material_Type": selected["Upper_Material_Type"],
                "Midsole": selected["Midsole"],
                "Outsole": selected["Outsole"],
                "Mass_kg": round(float(selected["Mass_kg"]), 3),
                "Price_USD": round(float(selected["Price_USD"]), 2),
                "Lifespan_Years": round(float(selected["Lifespan_Years"]), 2),
                "CO2e_Materials": round(material, 3),
                "CO2e_Manufacturing": round(float(selected["CO2e_Manufacturing"]), 3),
                "CO2e_Transport": round(float(selected["CO2e_Transport"]), 3),
                "CO2e_Packaging": round(float(selected["CO2e_Packaging"]), 3),
                "CO2e_EOL": round(float(selected["CO2e_EOL"]), 3),
                "CO2e_Total": round(float(selected["CO2e_Total"]), 3),
                "Annualized_CO2e": round(float(selected["Annualized_CO2e"]), 3),
                "Source_Type": source_row["Source_Type"],
                "Mass_Source_Type": classify_mass_source(source_row["Mass_Basis"]),
                "Mass_Basis": source_row["Mass_Basis"],
                "Lifespan_Basis": source_row["Lifespan_Basis"],
                "Downstream_Adder_Logic": "Manufacturing ~= 0.25M; Transport ~= 0.15M; Packaging ~= 0.05M; End-of-life ~= 0.02M",
                "Formula_Summary": (
                    f"Materials {selected['CO2e_Materials']:.2f} + Manufacturing {selected['CO2e_Manufacturing']:.2f} "
                    f"+ Transport {selected['CO2e_Transport']:.2f} + Packaging {selected['CO2e_Packaging']:.2f} "
                    f"+ EOL {selected['CO2e_EOL']:.2f} = Total {selected['CO2e_Total']:.2f} kg CO2e."
                ),
            }
        )
    return pd.DataFrame(rows)


def assumption_rules(df: pd.DataFrame, dataset_name: str, product_sources_name: str, outdir_name: str) -> pd.DataFrame:
    lifespan_counts = df["Lifespan_Years"].value_counts().sort_index()
    rows = [
        {
            "Assumption_Group": "Functional unit",
            "Code_or_Bucket": "Per pair",
            "Operational_Definition": "Primary comparison unit is one reference pair of adult footwear at purchase point; durability is analyzed separately through Annualized_CO2e.",
            "Observed_Count": len(df),
            "Released_Evidence": f"{dataset_name}; {outdir_name}/S19_Functional_Unit_Sensitivity.csv",
        },
        {
            "Assumption_Group": "Size normalization",
            "Code_or_Bucket": "Reference adult product mass proxy",
            "Operational_Definition": "The screening model does not parameterize exact shoe size. Mass_kg is treated as a stored reference mass proxy suitable for cross-product comparison rather than as a directly measured pair weight.",
            "Observed_Count": len(df),
            "Released_Evidence": dataset_name,
        },
        {
            "Assumption_Group": "Mass coding",
            "Code_or_Bucket": "Stored mass proxy in kilograms",
            "Operational_Definition": "Public-disclosure validation suggests the released Mass_kg field aligns more closely with single-shoe-scale official weights than with doubled pair-converted mass. Where no disclosure is available, the stored value is inferred from comparable products with similar category and upper/midsole/outsole architecture. The published 40-row carbon totals are not retroactively row-scaled from this diagnostic subset.",
            "Observed_Count": len(df),
            "Released_Evidence": f"{product_sources_name}; {outdir_name}/S14_Inventory_Calculation.csv; public_weight_validation_subset.csv",
        },
        {
            "Assumption_Group": "Price coding",
            "Code_or_Bucket": "Public list price",
            "Operational_Definition": "Price_USD represents the public list price used in the March 2026 screening dataset and may differ from later markdowns or regional colorway prices on reconstructed pages.",
            "Observed_Count": len(df),
            "Released_Evidence": f"{dataset_name}; {product_sources_name}",
        },
        {
            "Assumption_Group": "Source provenance",
            "Code_or_Bucket": "Official page hierarchy",
            "Operational_Definition": "Exact official PDPs were preferred; when unrecoverable, official collection pages, regional PDPs, or official release pages were used and labeled explicitly in the released appendix.",
            "Observed_Count": len(df),
            "Released_Evidence": f"{product_sources_name}; {outdir_name}/S20_Provenance_Summary.csv",
        },
    ]
    lifespan_definitions = {
        "1.0": "Daily-use screening assumption for high-mileage performance footwear.",
        "1.5": "Moderate-use screening assumption for versatile athletic footwear or lighter lifestyle/cleated constructions.",
        "2.0": "Longer-life screening assumption for sturdier performance or partial-/full-leather constructions.",
        "2.5": "Extended casual-use screening assumption for durable lifestyle sneakers.",
        "3.0": "Multi-year casual-use screening assumption for classic leather sneakers.",
        "5.0": "Long-life screening assumption for heavy leather oxford/boot-like construction.",
    }
    for bucket, definition in lifespan_definitions.items():
        rows.append(
            {
                "Assumption_Group": "Lifespan coding",
                "Code_or_Bucket": bucket,
                "Operational_Definition": definition,
                "Observed_Count": int(lifespan_counts.get(float(bucket), 0) or lifespan_counts.get(bucket, 0)),
                "Released_Evidence": f"{dataset_name}; {product_sources_name}",
            }
        )
    return pd.DataFrame(rows)


def sampling_sensitivity(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    categories = ["Running", "Basketball", "Casual", "Soccer"]
    samples = [("All_Categories", df, "Baseline screening sample")]
    samples.extend(
        [
            (f"Drop_{category}", df.loc[df["Category"] != category].reset_index(drop=True), f"Leave-one-category-out sensitivity excluding {category}")
            for category in categories
        ]
    )
    for label, sample, note in samples:
        X = pd.DataFrame({"Intercept": 1.0, "Price_USD": sample["Price_USD"], "Is_Leather": sample["Is_Leather"]})
        result = ols(sample["CO2e_Total"], X)
        leather_row = result.coefficients.loc["Is_Leather"]
        rows.append(
            {
                "Analysis": label,
                "Type": "Leave_One_Category_Out" if label != "All_Categories" else "Baseline",
                "N": len(sample),
                "Leather_N": int(sample["Is_Leather"].sum()),
                "NonLeather_N": int((1 - sample["Is_Leather"]).sum()),
                "Leather_Estimate": round(float(leather_row["coef"]), 6),
                "Leather_SE": round(float(leather_row["se"]), 6),
                "Leather_P_Value": float(leather_row["p"]),
                "R2": round(result.r2, 6),
                "Note": note,
            }
        )

    rng = np.random.default_rng(BOOTSTRAP_SEED)
    coefficients = []
    leather = df.loc[df["Is_Leather"] == 1]
    nonleather = df.loc[df["Is_Leather"] == 0]
    for _ in range(2_000):
        sample = pd.concat(
            [
                leather.iloc[rng.integers(0, len(leather), len(leather))],
                nonleather.iloc[rng.integers(0, len(nonleather), len(leather))],
            ],
            ignore_index=True,
        )
        X = pd.DataFrame({"Intercept": 1.0, "Price_USD": sample["Price_USD"], "Is_Leather": sample["Is_Leather"]})
        coefficients.append(float(ols(sample["CO2e_Total"], X).coefficients.loc["Is_Leather", "coef"]))
    coef_series = pd.Series(coefficients)
    rows.append(
        {
            "Analysis": "Balanced_Material_Bootstrap",
            "Type": "Resampled_Balanced_Material",
            "N": len(leather) * 2,
            "Leather_N": len(leather),
            "NonLeather_N": len(leather),
            "Leather_Estimate": round(float(coef_series.mean()), 6),
            "Leather_SE": round(float(coef_series.std(ddof=1)), 6),
            "Leather_P_Value": np.nan,
            "R2": np.nan,
            "Note": (
                f"2,000 resamples with equal leather/non-leather counts; 95% interval "
                f"[{coef_series.quantile(0.025):.3f}, {coef_series.quantile(0.975):.3f}]"
            ),
        }
    )
    return pd.DataFrame(rows)


def market_weighted_leather_scenarios(df: pd.DataFrame) -> pd.DataFrame:
    rng = np.random.default_rng(BOOTSTRAP_SEED)
    rows = []

    category_pools: dict[tuple[str, int], pd.DataFrame] = {}
    for category in ["Running", "Basketball", "Casual", "Soccer"]:
        for is_leather in [0, 1]:
            category_pools[(category, is_leather)] = df.loc[
                (df["Category"] == category) & (df["Is_Leather"] == is_leather)
            ].reset_index(drop=True)

    scenario_defs = [
        ("Observed_Casual_Leather_80pct", 8, 2, "Observed released sample composition."),
        (
            "Scenario_Casual_Leather_60pct",
            6,
            4,
            "Hypothetical market-weighted scenario with casual leather share reduced from 80% to 60% while other category counts are held fixed.",
        ),
        (
            "Scenario_Casual_Leather_50pct",
            5,
            5,
            "Hypothetical market-weighted scenario with a balanced 50/50 casual leather split.",
        ),
        (
            "Scenario_Casual_Leather_40pct",
            4,
            6,
            "Hypothetical market-weighted scenario with casual leather share reduced to 40%.",
        ),
        (
            "Scenario_Casual_Leather_30pct",
            3,
            7,
            "Hypothetical market-weighted scenario with casual leather share reduced to 30%.",
        ),
        (
            "Scenario_Casual_Leather_20pct",
            2,
            8,
            "Hypothetical market-weighted scenario with casual leather share reduced to 20%.",
        ),
    ]

    for label, casual_leather_n, casual_nonleather_n, note in scenario_defs:
        coefficients = []
        for _ in range(2_000):
            sample_parts = []
            sample_plan = {
                ("Running", 0): 10,
                ("Running", 1): 0,
                ("Basketball", 0): 10,
                ("Basketball", 1): 0,
                ("Soccer", 0): 6,
                ("Soccer", 1): 4,
                ("Casual", 0): casual_nonleather_n,
                ("Casual", 1): casual_leather_n,
            }
            for key, n in sample_plan.items():
                if n == 0:
                    continue
                pool = category_pools[key]
                picks = rng.integers(0, len(pool), n)
                sample_parts.append(pool.iloc[picks])
            sample = pd.concat(sample_parts, ignore_index=True)
            X = pd.DataFrame({"Intercept": 1.0, "Price_USD": sample["Price_USD"], "Is_Leather": sample["Is_Leather"]})
            coefficients.append(float(ols(sample["CO2e_Total"], X).coefficients.loc["Is_Leather", "coef"]))

        coef_series = pd.Series(coefficients)
        total_leather = casual_leather_n + 4
        total_nonleather = 40 - total_leather
        rows.append(
            {
                "Scenario": label,
                "Casual_Leather_N": casual_leather_n,
                "Casual_NonLeather_N": casual_nonleather_n,
                "Overall_Leather_N": total_leather,
                "Overall_Leather_Share": round(total_leather / 40.0, 3),
                "Resamples": 2000,
                "Leather_Estimate_Mean": round(float(coef_series.mean()), 6),
                "Leather_Estimate_SE": round(float(coef_series.std(ddof=1)), 6),
                "CI_2.5": round(float(coef_series.quantile(0.025)), 6),
                "CI_50": round(float(coef_series.quantile(0.5)), 6),
                "CI_97.5": round(float(coef_series.quantile(0.975)), 6),
                "Note": note,
            }
        )
    return pd.DataFrame(rows)


def phase_split_scenarios(df: pd.DataFrame) -> pd.DataFrame:
    base_total = df["CO2e_Total"].copy()

    def scenario_totals(frame: pd.DataFrame, name: str) -> pd.Series:
        if name == "Canonical_68_17_10_3_2":
            shares = {"materials": 0.68, "manufacturing": 0.17, "transport": 0.10, "packaging": 0.03, "eol": 0.02}
            total = frame["CO2e_Materials"] / shares["materials"]
            return total
        if name == "Cheah_2013_Manufacturing_Dominant":
            # Cheah et al. report ~29% materials and ~68% manufacturing for a synthetic running shoe.
            # The residual 3% is distributed across transport/packaging/end-of-life in the canonical 10:3:2 ratio.
            shares = {"materials": 0.29, "manufacturing": 0.68, "transport": 0.02, "packaging": 0.006, "eol": 0.004}
            total = frame["CO2e_Materials"] / shares["materials"]
            return total
        if name == "Bodoga_2024_Component_Dominant":
            # Bodoga et al. report a professional-footwear case with five-stage shares explicitly reported.
            shares = {"materials": 0.7981, "manufacturing": 0.0703, "transport": 0.0278, "packaging": 0.0237, "eol": 0.0802}
            total = frame["CO2e_Materials"] / shares["materials"]
            return total
        if name == "Hybrid_Cheah_NonLeather_Bodoga_Leather":
            totals = []
            for _, row in frame.iterrows():
                material_share = 0.7981 if int(row["Is_Leather"]) == 1 else 0.29
                totals.append(float(row["CO2e_Materials"]) / material_share)
            return pd.Series(totals, index=frame.index)
        raise ValueError(name)

    scenario_defs = [
        (
            "Canonical_68_17_10_3_2",
            "Current manuscript screening split with materials fixed at 68% of total.",
        ),
        (
            "Cheah_2013_Manufacturing_Dominant",
            "Literature-informed manufacturing-dominant split based on Cheah et al. 2013 (materials ~29%, manufacturing ~68%, residual 3% distributed proportionally across downstream minor stages).",
        ),
        (
            "Bodoga_2024_Component_Dominant",
            "Literature-informed component-dominant split based on Bodoga et al. 2024 (materials 79.81%, manufacturing 7.03%, packaging 2.37%, distribution 2.78%, end-of-life 8.02%).",
        ),
        (
            "Hybrid_Cheah_NonLeather_Bodoga_Leather",
            "Material-informed hybrid scenario using the Cheah manufacturing-dominant split for non-leather rows and the Bodoga component-dominant split for leather-associated rows.",
        ),
    ]

    rows = []
    for name, note in scenario_defs:
        total = scenario_totals(df, name)
        scenario_df = df.copy()
        scenario_df["Scenario_CO2e_Total"] = total
        X = pd.DataFrame({"Intercept": 1.0, "Price_USD": scenario_df["Price_USD"], "Is_Leather": scenario_df["Is_Leather"]})
        result = ols(scenario_df["Scenario_CO2e_Total"], X)
        leather_row = result.coefficients.loc["Is_Leather"]
        rows.append(
            {
                "Scenario": name,
                "Mean_CO2e": round(float(total.mean()), 3),
                "Median_CO2e": round(float(total.median()), 3),
                "Leather_Estimate": round(float(leather_row["coef"]), 6),
                "Leather_SE": round(float(leather_row["se"]), 6),
                "Leather_P_Value": float(leather_row["p"]),
                "R2": round(float(result.r2), 6),
                "Spearman_vs_Canonical": round(float(stats.spearmanr(base_total, total).statistic), 6),
                "Note": note,
            }
        )
    return pd.DataFrame(rows)


def emission_factor_provenance() -> pd.DataFrame:
    rows = [
        {
            "Material": "Bovine leather",
            "Selected_GWP_kgCO2e_per_kg": 25.0,
            "Selected_Water_L_per_kg": 17100,
            "Selected_Ecotox_CTUe_per_kg": 410,
            "Primary_Source_Family": "Cascale Higg MSI / ecoinvent-informed screening proxy",
            "Primary_Source_Detail": "Chrome-tanned bovine leather family used as the central screening factor; exact licensed dataset records are not redistributed in the manuscript package.",
            "Public_Benchmark_1": "Milà i Canals et al. 2002: leather supply-chain burdens are material enough to drive eco-label criteria.",
            "Public_Benchmark_2": "Yang et al. 2021: chrome-tanned cowhide upper-leather processing alone is ~7.04 kg CO2e/kg, indicating that upstream-inclusive leather factors should exceed tannery-only burdens.",
            "Scenario_Low": 18.0,
            "Scenario_Central": 25.0,
            "Scenario_High": 32.0,
            "Interpretation": "Chosen as an upstream-inclusive central screening factor rather than a tannery-only processing factor.",
        },
        {
            "Material": "Synthetic leather",
            "Selected_GWP_kgCO2e_per_kg": 8.0,
            "Selected_Water_L_per_kg": 85,
            "Selected_Ecotox_CTUe_per_kg": 45,
            "Primary_Source_Family": "Cascale Higg MSI / ecoinvent-informed coated-synthetic proxy",
            "Primary_Source_Detail": "Used as a directional PU/PVC-coated upper-material screening proxy for footwear uppers.",
            "Public_Benchmark_1": "Screening proxy anchored to coated-textile and synthetic-upper families used in apparel-footwear LCAs.",
            "Public_Benchmark_2": "Monte Carlo and tornado sensitivity treat this factor as one of the key uncertain material inputs.",
            "Scenario_Low": 6.0,
            "Scenario_Central": 8.0,
            "Scenario_High": 10.0,
            "Interpretation": "Reported as a central coated-synthetic screening factor, not as a single audited supplier-specific record.",
        },
        {
            "Material": "Polyester textile",
            "Selected_GWP_kgCO2e_per_kg": 5.5,
            "Selected_Water_L_per_kg": 71,
            "Selected_Ecotox_CTUe_per_kg": 38,
            "Primary_Source_Family": "Cascale Higg MSI / ecoinvent polyester textile proxy",
            "Primary_Source_Detail": "Used as the central fossil-based textile upper proxy.",
            "Public_Benchmark_1": "Aligned to widely used synthetic-textile screening factors in apparel LCA practice.",
            "Public_Benchmark_2": "Scenario band captures process-energy and fabric-construction variability without implying a single universal polyester value.",
            "Scenario_Low": 4.0,
            "Scenario_Central": 5.5,
            "Scenario_High": 7.0,
            "Interpretation": "Central textile factor used for comparable footwear uppers rather than a fabric-style-specific declaration.",
        },
        {
            "Material": "Recycled polyester",
            "Selected_GWP_kgCO2e_per_kg": 2.8,
            "Selected_Water_L_per_kg": 32,
            "Selected_Ecotox_CTUe_per_kg": 18,
            "Primary_Source_Family": "Cascale Higg MSI recycled-polyester proxy",
            "Primary_Source_Detail": "Directional recycled-PET textile factor used for lower-impact synthetic uppers.",
            "Public_Benchmark_1": "Selected value is below the fossil polyester proxy and preserves the expected recycled-versus-virgin ordering.",
            "Public_Benchmark_2": "Scenario band reflects recycling-route and electricity-mix sensitivity.",
            "Scenario_Low": 2.0,
            "Scenario_Central": 2.8,
            "Scenario_High": 3.6,
            "Interpretation": "Used as a recycled-content screening proxy, not as a brand-specific claim.",
        },
        {
            "Material": "EVA foam",
            "Selected_GWP_kgCO2e_per_kg": 3.1,
            "Selected_Water_L_per_kg": 45,
            "Selected_Ecotox_CTUe_per_kg": 22,
            "Primary_Source_Family": "ecoinvent-informed EVA polymer proxy",
            "Primary_Source_Detail": "Used for midsole-dominant athletic footwear architectures.",
            "Public_Benchmark_1": "Aligned to polymer-foam screening factors commonly used in footwear and sporting-goods LCA work.",
            "Public_Benchmark_2": "Scenario band reflects resin and process-energy uncertainty rather than a single supplier-specific formulation.",
            "Scenario_Low": 2.5,
            "Scenario_Central": 3.1,
            "Scenario_High": 4.0,
            "Interpretation": "Central midsole screening factor for athletic footwear, not a chemistry-specific audited recipe.",
        },
        {
            "Material": "Rubber",
            "Selected_GWP_kgCO2e_per_kg": 3.0,
            "Selected_Water_L_per_kg": 65,
            "Selected_Ecotox_CTUe_per_kg": 28,
            "Primary_Source_Family": "ecoinvent-informed rubber proxy",
            "Primary_Source_Detail": "Directional outsole-material factor for mixed synthetic/natural rubber use in footwear.",
            "Public_Benchmark_1": "Consistent with moderate-burden outsole screening assumptions relative to leather and polyester uppers.",
            "Public_Benchmark_2": "Scenario band reflects compound-specific and natural-versus-synthetic mix uncertainty.",
            "Scenario_Low": 2.4,
            "Scenario_Central": 3.0,
            "Scenario_High": 3.8,
            "Interpretation": "Central outsole factor for mixed rubber constructions.",
        },
        {
            "Material": "Cotton canvas",
            "Selected_GWP_kgCO2e_per_kg": 4.5,
            "Selected_Water_L_per_kg": 10000,
            "Selected_Ecotox_CTUe_per_kg": 55,
            "Primary_Source_Family": "Cascale Higg MSI / Water Footprint Network cotton-textile proxy",
            "Primary_Source_Detail": "Directional cotton-canvas upper factor paired with high blue/green-water burden expectations.",
            "Public_Benchmark_1": "Water burden is intentionally high relative to synthetics, consistent with cotton irrigation literature.",
            "Public_Benchmark_2": "Scenario band reflects cultivation-region variability and textile finishing uncertainty.",
            "Scenario_Low": 3.5,
            "Scenario_Central": 4.5,
            "Scenario_High": 6.0,
            "Interpretation": "Used as a cotton upper screening proxy with high water sensitivity.",
        },
    ]
    return pd.DataFrame(rows)


def functional_unit_sensitivity(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    pair_based = df["CO2e_Total"]
    for outcome, label in [("CO2e_Total", "Per_Pair_Primary"), ("Annualized_CO2e", "Annualized_Sensitivity")]:
        y = df[outcome]
        X = pd.DataFrame({"Intercept": 1.0, "Price_USD": df["Price_USD"], "Is_Leather": df["Is_Leather"]})
        result = ols(y, X)
        means = df.groupby("Leather_Group")[outcome].mean()
        leather_mean = float(means["Leather-associated"])
        nonleather_mean = float(means["Non-leather"])
        rows.append(
            {
                "Outcome": label,
                "Definition": "One pair of footwear" if outcome == "CO2e_Total" else "CO2e_Total / Lifespan_Years",
                "Leather_Mean": round(leather_mean, 6),
                "NonLeather_Mean": round(nonleather_mean, 6),
                "Leather_to_NonLeather_Ratio": round(leather_mean / nonleather_mean, 6),
                "Leather_Estimate": round(float(result.coefficients.loc["Is_Leather", "coef"]), 6),
                "Leather_SE": round(float(result.coefficients.loc["Is_Leather", "se"]), 6),
                "Leather_P_Value": float(result.coefficients.loc["Is_Leather", "p"]),
                "R2": round(result.r2, 6),
                "Spearman_vs_Per_Pair": round(float(stats.spearmanr(pair_based, y).statistic), 6),
            }
        )
    return pd.DataFrame(rows)


def provenance_summary(product_sources: pd.DataFrame) -> pd.DataFrame:
    source_definitions = {
        "official_exact_pdp": "Exact official product-detail page recovered for the sampled model or a directly comparable colorway/SKU.",
        "official_exact_pdp_redirect": "Direct official SKU URL recovered, but the current request resolves to a broader official collection or landing page.",
        "official_collection_page": "Official model-family or collection page used when a stable single exact PDP was not recoverable.",
        "official_regional_pdp": "Official regional product-detail page used when a comparable U.S. PDP was not recoverable.",
        "official_release_page": "Official brand release or newsroom page used when it preserved launch features and MSRP but not a stable PDP.",
    }
    rows = []
    total = len(product_sources)
    counts = product_sources["Source_Type"].value_counts()
    for source_type, definition in source_definitions.items():
        count = int(counts.get(source_type, 0))
        rows.append(
            {
                "Metric": source_type,
                "Value": count,
                "Share_of_Sample": round(count / total, 6),
                "Definition_or_Interpretation": definition,
            }
        )
    exact_like = int(counts.get("official_exact_pdp", 0))
    proxy_count = int(total - exact_like)
    rows.append(
        {
            "Metric": "proxy_or_non_us_page_used",
            "Value": proxy_count,
            "Share_of_Sample": round(proxy_count / total, 6),
            "Definition_or_Interpretation": "Collection, regional, or release pages were used as transparent proxies when exact U.S. PDPs were not stably recoverable.",
        }
    )

    local_capture_cols = [
        "Local_HTML_Artifact",
        "Local_PDF_Artifact",
        "Local_Screenshot_Artifact",
    ]
    if all(col in product_sources.columns for col in local_capture_cols):
        local_capture_complete = int(
            product_sources[local_capture_cols]
            .fillna("")
            .apply(lambda col: col.astype(str).str.strip())
            .ne("")
            .all(axis=1)
            .sum()
        )
        rows.append(
            {
                "Metric": "local_capture_complete_rows",
                "Value": local_capture_complete,
                "Share_of_Sample": round(local_capture_complete / total, 6),
                "Definition_or_Interpretation": "Rows with local HTML, PDF, and screenshot evidence stored in archived_specs for the Phase 1 provenance-hardening pass.",
            }
        )

    archive_url_count = 0
    if "Archive_URL" in product_sources.columns:
        archive_url_count = int(product_sources["Archive_URL"].fillna("").astype(str).str.strip().ne("").sum())
    rows.append(
        {
            "Metric": "external_archive_url_rows",
            "Value": archive_url_count,
            "Share_of_Sample": round(archive_url_count / total, 6),
            "Definition_or_Interpretation": "Rows with a populated third-party archive URL in addition to the local evidence captures.",
        }
    )
    return pd.DataFrame(rows)


def expansion_comparison() -> pd.DataFrame:
    canonical_path = ROOT / "expanded_dataset_40_models.csv"
    resolved_path = ROOT / "expanded_dataset_60_resolved.csv"
    canonical_sources_path = ROOT / "product_source_appendix_40_models.csv"
    resolved_sources_path = ROOT / "product_source_appendix_60_resolved.csv"
    if not (canonical_path.exists() and resolved_path.exists() and canonical_sources_path.exists() and resolved_sources_path.exists()):
        return pd.DataFrame(
            [
                {
                    "Dataset": "Unavailable",
                    "Rows": np.nan,
                    "Leather_Count": np.nan,
                    "NonLeather_Count": np.nan,
                    "Mean_CO2e": np.nan,
                    "Median_CO2e": np.nan,
                    "Leather_Estimate": np.nan,
                    "R2": np.nan,
                    "Exact_PDP_Share": np.nan,
                    "Proxy_Page_Share": np.nan,
                    "Interpretation": "Resolved-source extension files not available for comparison.",
                }
            ]
        )

    def summarize(dataset_path: Path, sources_path: Path, label: str) -> dict[str, object]:
        data = load_dataset(dataset_path)
        sources = pd.read_csv(sources_path)
        X = pd.DataFrame({"Intercept": 1.0, "Price_USD": data["Price_USD"], "Is_Leather": data["Is_Leather"]})
        result = ols(data["CO2e_Total"], X)
        exact = (sources["Source_Type"] == "official_exact_pdp").mean()
        proxy = 1.0 - exact
        return {
            "Dataset": label,
            "Rows": len(data),
            "Leather_Count": int(data["Is_Leather"].sum()),
            "NonLeather_Count": int((1 - data["Is_Leather"]).sum()),
            "Mean_CO2e": round(float(data["CO2e_Total"].mean()), 3),
            "Median_CO2e": round(float(data["CO2e_Total"].median()), 3),
            "Leather_Estimate": round(float(result.coefficients.loc["Is_Leather", "coef"]), 3),
            "R2": round(float(result.r2), 3),
            "Exact_PDP_Share": round(float(exact), 3),
            "Proxy_Page_Share": round(float(proxy), 3),
            "Interpretation": "",
        }

    rows = [
        summarize(canonical_path, canonical_sources_path, "Canonical_40_Row"),
        summarize(resolved_path, resolved_sources_path, "Resolved_60_Row"),
    ]
    rows.append(
        {
            "Dataset": "Difference_60_minus_40",
            "Rows": rows[1]["Rows"] - rows[0]["Rows"],
            "Leather_Count": rows[1]["Leather_Count"] - rows[0]["Leather_Count"],
            "NonLeather_Count": rows[1]["NonLeather_Count"] - rows[0]["NonLeather_Count"],
            "Mean_CO2e": round(rows[1]["Mean_CO2e"] - rows[0]["Mean_CO2e"], 3),
            "Median_CO2e": round(rows[1]["Median_CO2e"] - rows[0]["Median_CO2e"], 3),
            "Leather_Estimate": round(rows[1]["Leather_Estimate"] - rows[0]["Leather_Estimate"], 3),
            "R2": round(rows[1]["R2"] - rows[0]["R2"], 3),
            "Exact_PDP_Share": round(rows[1]["Exact_PDP_Share"] - rows[0]["Exact_PDP_Share"], 3),
            "Proxy_Page_Share": round(rows[1]["Proxy_Page_Share"] - rows[0]["Proxy_Page_Share"], 3),
            "Interpretation": "The resolved-source extension is reported as a directional robustness layer, not as a replacement inferential baseline.",
        }
    )
    return pd.DataFrame(rows)


def literature_hotspot_context() -> pd.DataFrame:
    rows = [
        {
            "Study": "Cheah et al. 2013",
            "Footwear_Context": "Synthetic running shoe case study",
            "Main_Hotspot_Phase": "Manufacturing",
            "Interpretive_Use": "Shows that footwear hotspot phases can vary by product architecture and scope assumptions.",
        },
        {
            "Study": "Gottfridsson and Zhang 2015",
            "Footwear_Context": "Swedish shoe-consumption model",
            "Main_Hotspot_Phase": "Materials",
            "Interpretive_Use": "Supports the expectation that upstream material production can dominate comparative footwear impacts.",
        },
        {
            "Study": "Milà i Canals et al. 2002",
            "Footwear_Context": "Leather eco-label criteria framework",
            "Main_Hotspot_Phase": "Leather supply-chain burden",
            "Interpretive_Use": "Supports the emphasis on leather-associated upstream burdens rather than generic brand-level claims.",
        },
        {
            "Study": "Current manuscript",
            "Footwear_Context": "40-model comparative screening dataset",
            "Main_Hotspot_Phase": "Materials-stage estimate plus fixed downstream adders",
            "Interpretive_Use": "Tests comparative material effects directly instead of assuming a universal footwear phase hierarchy.",
        },
    ]
    return pd.DataFrame(rows)


def source_files(dataset_name: str, product_sources_name: str, outdir_name: str, workbook_name: str) -> pd.DataFrame:
    rows = [
        {
            "File": dataset_name,
            "Role": "Canonical product-level dataset used for the current analysis run",
        },
        {
            "File": product_sources_name,
            "Role": "Reconstructed model-level public source log with URLs, access dates, and coding notes aligned to the current dataset",
        },
        {
            "File": "water_footprint_analysis.md",
            "Role": "Supporting water-footprint interpretation aligned to the 40-model dataset",
        },
        {
            "File": "comparative_benchmarking.md",
            "Role": "Supporting wardrobe and apparel-context comparison",
        },
        {
            "File": "additional_sensitivity_analyses.md",
            "Role": "Supporting bootstrap, scenario, and Monte Carlo narrative",
        },
        {
            "File": "public_weight_validation_subset.csv",
            "Role": "Official brand disclosed-weight validation subset used to reconcile the stored Mass_kg field definition",
        },
        {
            "File": f"{outdir_name}/S24_Phase_Split_Scenarios.csv",
            "Role": "Literature-informed alternative phase-allocation scenarios and their effect on totals, rankings, and the leather coefficient",
        },
        {
            "File": f"{outdir_name}/S25_Emission_Factor_Provenance.csv",
            "Role": "Material-factor provenance sheet with source families, public benchmark notes, and low/central/high scenario bands",
        },
        {
            "File": f"{outdir_name}/",
            "Role": "Directory containing generated supplementary CSV outputs for the current analysis run",
        },
        {
            "File": workbook_name,
            "Role": "Workbook bundling the generated supplementary sheets for the current analysis run",
        },
        {
            "File": "tools/analyze_expanded_dataset.py",
            "Role": "Recreates supplementary statistical outputs from a chosen dataset/source appendix pair",
        },
        {
            "File": "tools/generate_figure_set.py",
            "Role": "Recreates the maintained manuscript and supplementary figures",
        },
        {
            "File": "tools/regenerate_artifacts.py",
            "Role": "Regenerates docx/pdf/xlsx artifacts from markdown and CSV sources",
        },
    ]
    return pd.DataFrame(rows)


def build_workbook(frames: dict[str, pd.DataFrame], workbook_path: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for title, frame in frames.items():
        dataframe_to_sheet(workbook, title, frame)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)


def main() -> None:
    args = parse_args()
    dataset_path = args.dataset.resolve()
    product_sources_path = args.product_sources.resolve()
    outdir = args.outdir.resolve()
    workbook = args.workbook.resolve()
    dataset_name = dataset_path.name
    product_sources_name = product_sources_path.name
    outdir_name = outdir.name
    workbook_name = workbook.name

    outdir.mkdir(parents=True, exist_ok=True)
    df = load_dataset(dataset_path)

    summary = summary_statistics(df)
    categories = category_statistics(df)
    materials = material_statistics(df)
    rules = reproducibility_rules(dataset_name, outdir_name)
    tests = correlations_and_tests(df)
    anova_frame, tukey_frame = anova_and_tukey(df)
    regressions = regression_models(df)
    leave_one_out_full, leave_one_out_summary = leave_one_out(df)
    bootstrap = bootstrap_intervals(df)
    sources = source_files(dataset_name, product_sources_name, outdir_name, workbook_name)
    product_sources = pd.read_csv(product_sources_path)
    inventory = inventory_calculation_sheet(df, product_sources)
    worked = worked_examples(df, product_sources)
    assumptions = assumption_rules(df, dataset_name, product_sources_name, outdir_name)
    sampling = sampling_sensitivity(df)
    market_weighted = market_weighted_leather_scenarios(df)
    diagnostics = regression_diagnostics(df)
    functional_unit = functional_unit_sensitivity(df)
    provenance = provenance_summary(product_sources)
    expansion = expansion_comparison()
    hotspot_context = literature_hotspot_context()
    phase_scenarios = phase_split_scenarios(df)
    factor_provenance = emission_factor_provenance()

    frames = {
        "S0_Source_Files": sources,
        "S1_Reproducibility": rules,
        "S2_Summary_Stats": summary,
        "S3_Category_Stats": categories,
        "S4_Material_Stats": materials,
        "S5_Tests": tests,
        "S6_ANOVA": anova_frame,
        "S7_Tukey": tukey_frame,
        "S8_Regressions": regressions,
        "S9_Leave_One_Out": leave_one_out_full,
        "S10_LOO_Summary": leave_one_out_summary,
        "S11_Bootstrap": bootstrap,
        "S12_Dataset": df,
        "S13_Product_Sources": product_sources,
        "S14_Inventory_Calculation": inventory,
        "S15_Worked_Examples": worked,
        "S16_Assumption_Rules": assumptions,
        "S17_Sampling_Sensitivity": sampling,
        "S18_Regression_Diagnostics": diagnostics,
        "S19_Functional_Unit_Sensitivity": functional_unit,
        "S20_Provenance_Summary": provenance,
        "S21_Expansion_Comparison": expansion,
        "S22_Literature_Hotspot_Context": hotspot_context,
        "S23_Market_Weighted_Leather_Scenarios": market_weighted,
        "S24_Phase_Split_Scenarios": phase_scenarios,
        "S25_Emission_Factor_Provenance": factor_provenance,
    }

    for title, frame in frames.items():
        frame.to_csv(outdir / f"{title}.csv", index=False)

    build_workbook(frames, workbook)


if __name__ == "__main__":
    main()
